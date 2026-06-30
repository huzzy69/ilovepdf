import openpyxl
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
import tempfile
import os

def convert_excel_to_pdf(excel_path: str) -> str:
    """
    Converts an Excel spreadsheet (xlsx) to a styled PDF using openpyxl and reportlab.
    """
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    fd, output_path = tempfile.mkstemp(suffix=".pdf")
    os.close(fd)
    
    doc = SimpleDocTemplate(
        output_path, 
        pagesize=landscape(letter), 
        rightMargin=30, 
        leftMargin=30, 
        topMargin=30, 
        bottomMargin=30
    )
    story = []
    styles = getSampleStyleSheet()
    title_style = styles['Heading1']
    cell_style = ParagraphStyle('CellStyle', parent=styles['Normal'], fontSize=8, leading=10)
    
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        story.append(Paragraph(f"Sheet: {sheet_name}", title_style))
        story.append(Spacer(1, 10))
        
        # Read cells
        data = []
        for row in sheet.iter_rows(values_only=True):
            # Ignore empty rows
            if all(val is None for val in row):
                continue
            row_data = []
            for val in row:
                val_str = "" if val is None else str(val)
                row_data.append(Paragraph(val_str, cell_style))
            data.append(row_data)
            
        if data:
            # Create ReportLab Table
            t = Table(data, hAlign='LEFT')
            t.setStyle(TableStyle([
                ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
                ('BACKGROUND', (0,0), (-1,0), colors.lightgrey),
                ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ]))
            story.append(t)
            story.append(Spacer(1, 20))
            
    doc.build(story)
    return output_path
